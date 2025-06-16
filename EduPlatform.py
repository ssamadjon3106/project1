# eduplatform.py (Refactored with CLI)

from abc import ABC, abstractmethod
from enum import Enum
from datetime import datetime
import hashlib
from typing import List, Dict
import csv
import openpyxl
from openpyxl import Workbook
from datetime import datetime

class UserRole(Enum):
    ADMIN = "admin"
    TEACHER = "teacher"
    STUDENT = "student"
    PARENT = "parent"

class AbstractRole(ABC):
    def __init__(self, full_name: str, email: str, password: str, role: UserRole):
        self._id = id(self)
        self.role = role
        self._full_name = full_name
        if "@" not in email:
            raise ValueError("Invalid email format")
        self._email = email
        self._password_hash = hashlib.sha256(password.encode()).hexdigest()
        self._created_at = datetime.now().isoformat()

    @abstractmethod
    def get_profile(self) -> dict:
        pass

    @abstractmethod
    def update_profile(self, **kwargs):
        pass

class Notification:
    def __init__(self, notif_id, message, sender_id, recipient_id):
        self.id = notif_id
        self.message = message
        self.sender_id = sender_id
        self.recipient_id = recipient_id
        self.created_at = datetime.now().isoformat()
        self.is_read = False

    def send(self, users: list):
        """Send this notification to the recipient's notification list"""
        recipient = next((u for u in users if u._id == self.recipient_id), None)
        if recipient:
            recipient.add_notification(self)
            return True
        return False

    def mark_as_read(self):
        self.is_read = True

    def __str__(self) -> str:  # Remove users parameter
        status = "Read" if self.is_read else "Unread"
        return (f"ID: {self.id} | Status: {status} | {self.created_at[:10]}\n"
                f"Message: {self.message}")
    


class User(AbstractRole):
    def __init__(self, full_name: str, email: str, password: str, role: UserRole):
        super().__init__(full_name, email, password, role)
        self._notifications = []

    def get_profile(self) -> dict:
        return {
            'id': self._id,
            'full_name': self._full_name,
            'email': self._email,
            'role': self.role.value,
            'created_at': self._created_at
        }

    def update_profile(self, **kwargs):
        for key, value in kwargs.items():
            if key == 'password':
                self._password_hash = hashlib.sha256(value.encode()).hexdigest()
            elif hasattr(self, f"_{key}"):
                setattr(self, f"_{key}", value)

    def add_notification(self, notification: Notification):
        self._notifications.append(notification)

    def view_notifications(self) -> list:
        return self._notifications

    def delete_notification(self, notif_id: int) -> bool:
        for i, note in enumerate(self._notifications):
            if note['id'] == notif_id:
                self._notifications.pop(i)
                return True
        return False
    def link_parent_child(self, parent_id: int, student_id: int, users: list) -> bool:
        parent = next((u for u in users if isinstance(u, Parent) and u._id == parent_id), None)
        student = next((u for u in users if isinstance(u, Student) and u._id == student_id), None)
        
        if not parent or not student:
            print("Parent or student not found")
            return False
        
        parent.add_child(student)
        print(f"Student {student._full_name} added to parent {parent._full_name}")
        return True

class Student(User):
    def __init__(self, full_name: str, email: str, password: str, grade: str):
        super().__init__(full_name, email, password, UserRole.STUDENT)
        self.grade = grade
        self.subjects = {}
        self.assignments = {}
        self.grades = {}

    def submit_assignment(self, assignment_id: int, content: str):
        if assignment_id in self.assignments:
         print("Already submitted.")
        else:
            self.assignments[assignment_id] = {
                "content": content,
                "submitted_at": datetime.now(),
                "grade": None,
                "comment": None,
            }
            print("Assignment submitted.")

    

    def view_grades(self):
        if not self.assignments:
            print("No assignments found.")
            return
        for aid, data in self.assignments.items():
            grade = data.get("grade", "Not graded")
            comment = data.get("comment", "")
            print(f"Assignment ID: {aid}")
            print(f"  Grade: {grade}")
            print(f"  Comment: {comment}")
            


    def calculate_average_grade(self) -> float:
        graded_assignments = [g['grade'] for g in self.assignments.values() 
                            if 'grade' in g and isinstance(g['grade'], (int, float))]
        if not graded_assignments:
            print('No graded assignments found')
            return 0.0
        average = sum(graded_assignments)/len(graded_assignments)    
        print(f'Average: {average:.2f}')
        return average
class Teacher(User):
    def __init__(self, full_name: str, email: str, password: str, subjects: List[str]):
        super().__init__(full_name, email, password, UserRole.TEACHER)
        self.subjects = subjects
        self.classes = []
        self._assignments = {}

    def create_assignment(self, title, description, deadline, subject, class_id):
        assignment = Assignment(title, description, deadline, subject, self._id, class_id)
        self._assignments[assignment.id] = assignment
        return assignment.id

    def grade_assignment(self, assignment_id: int, student_id: int, grade: int, comment: str, users: list):
        student = next((s for s in users if isinstance(s, Student) and s._id == student_id), None)
        if student and assignment_id in student.assignments:
            student.assignments[assignment_id]["grade"] = grade
            student.assignments[assignment_id]["comment"] = comment
            print("Graded successfully.")
            return True
        else:
            print("Failed to grade.")
            return False

    def send_to_parent(self, student_id: int, message: str, users: list):
        """Send a message to the parent(s) of a specific student"""
        # Find the student
        student = next((s for s in users if isinstance(s, Student) and s._id == student_id), None)
        if not student:
            print("Student not found.")
            return False
        
        # Find all parents of this student
        parents = [p for p in users if isinstance(p, Parent) and student_id in p.children]
        if not parents:
            print("No parents found for this student.")
            return False
        
        # Send notification to each parent
        success = True
        for parent in parents:
            notif_id = len(parent._notifications) + 1
            notification = Notification(notif_id, message, self._id, parent._id)
            if notification.send(users):
                print(f"Notification sent to parent {parent._full_name}")
            else:
                print(f"Failed to send to parent {parent._full_name}")
                success = False
        return success    



    def view_student_progress(self, student_id: int, users:list):
        student=next((u for u in users if isinstance(u, Student) and u._id==student_id), None)
        if student:
            if not self.assignments:
                print('NO assignments found')
            for aid, data in student.assignments.items():
                grade=data.get('grade', 'Not graded')
                comment=data.get('comment' , "")
                print(f'Assignment ID: {aid}')
                print(f' GRADE : {grade}')
                print(f' Comment: {comment}')
        else:
            print(f'Student with {student_id} not found')        



class Parent(User):
    def __init__(self, full_name: str, email: str, password: str):
        super().__init__(full_name, email, password, UserRole.PARENT)
        self.children = {}

    def add_child(self, student: Student):
        self.children[student._id] = student

    def view_child_grades(self, child_id: int , users: list):
        child = next((u for u in users if isinstance(u, Student) and u._id==child_id), None)
        if child: 
            if not child.assignments:
                print("No assignments found")

            for aid, data in child.assignments.items():
                grade = data.get("grade", "Not graded")
                comment = data.get("comment", "")
                print(f"Assignment ID: {aid}")
                print(f"  Grade: {grade}")
                print(f"  Comment: {comment}")
        else:
            print('child not found')
    def view_child_assignments(self, child_id, users: list):
        child = next((u for u in users if isinstance(u, Student) and u._id==child_id), None)
        if child: 
            if not child.assignments:
                print("No assignments found")
                return
            for aid, data in child.assignments.items():
                print(f"Assignment ID: {aid}")
                print(f"  Content: {data.get('content', '')}")
                print(f"  Grade: {data.get('grade', 'Not graded')}")
                print(f"  Comment: {data.get('comment', '')}")
                print()
        else:
            print("Child not found.")

    def send_to_parent(self, student_id: int, message: str, users: list):  # Fixed typo in method name
        student = next((u for u in users if isinstance(u, Student) and u._id == student_id), None)  
        if not student:
            print("Student not found")
            return False

        parent = next((p for p in users if isinstance(p, Parent) and student_id in p.children), None)
        if not parent:
            print("Parent not found for this student")
            return False

        notif_id = len(parent._notifications) + 1
        notification = Notification(notif_id, message, self._id, parent._id)

        if notification.send(users):
            print("Notification sent to parent successfully.")
            return True
        else:
            print("Failed to send notification.")
            return False
        

class Schedule:
    def __init__(self, schedule_id: int, class_id: str, day: str):
        self.schedule_id = schedule_id
        self.class_id = class_id
        self.day = day
        self.lessons = {}  # Format: {time: {'subject': str, 'teacher_id': int}}

    def add_lesson(self, time: str, subject: str, teacher_id: int, users: list) -> bool:
        """Add a lesson to the schedule"""
        # Check if teacher exists
        teacher = next((t for t in users if isinstance(t, Teacher) and t._id == teacher_id), None)
        if not teacher:
            print("Teacher not found.")
            return False
        
        # Check if time slot is available
        if time in self.lessons:
            print(f"Time slot {time} already occupied.")
            return False
        
        self.lessons[time] = {
            'subject': subject,
            'teacher_id': teacher_id
        }
        print(f"Lesson added at {time}: {subject} with teacher {teacher._full_name}")
        return True

    def remove_lesson(self, time: str) -> bool:
        """Remove a lesson from the schedule"""
        if time in self.lessons:
            del self.lessons[time]
            print(f"Lesson at {time} removed successfully.")
            return True
        print(f"No lesson found at {time}")
        return False

    def view_schedule(self, users: list) -> None:
        """Display the schedule with teacher names"""
        if not self.lessons:
            print("No lessons scheduled for this day.")
            return
        
        print(f"\nSchedule for Class {self.class_id} on {self.day}:")
        print("-" * 50)
        print("Time\t\tSubject\t\tTeacher")
        print("-" * 50)
        for time, lesson in sorted(self.lessons.items()):
            teacher = next((t for t in users if t._id == lesson['teacher_id']), None)
            teacher_name = teacher._full_name if teacher else "Unknown"
            print(f"{time}\t\t{lesson['subject']}\t\t{teacher_name}")


class Admin(User):
    
    def __init__(self, full_name: str, email: str, password: str):
        super().__init__(full_name, email, password, UserRole.ADMIN)
        self.permissions = []
        self.users = {}
        self.schedules = {}  # Format: {class_id: {day: Schedule}}
    def add_user(self, user: User, users: list) -> bool:
        if user._id in self.users:
            return False
        self.users[user._id] = user
        users.append(user)
        return True

    def remove_user(self, user_id: int, users: list) -> bool:
        user = self.users.pop(user_id, None)
        if user:
            users.remove(user)
            return True
        return False

    def create_schedule(self, class_id: str, day: str) -> bool:
        """Create a new schedule for a class"""
        if class_id not in self.schedules:
            self.schedules[class_id] = {}
        
        if day in self.schedules[class_id]:
            print(f"Schedule already exists for class {class_id} on {day}")
            return False
        
        schedule_id = len(self.schedules) + 1
        self.schedules[class_id][day] = Schedule(schedule_id, class_id, day)
        print(f"Schedule created for class {class_id} on {day}")
        return True

    def get_schedule(self, class_id: str, day: str) -> Schedule:
        """Get a schedule for a class on a specific day"""
        return self.schedules.get(class_id, {}).get(day)
    

class Assignment:
    def __init__(self, title, description, deadline, subject, teacher_id, class_id):
        self.id = id(self)
        self.title = title
        self.description = description
        self.deadline = deadline
        self.subject = subject
        self.teacher_id = teacher_id
        self.class_id = class_id
        self.submissions = {}
        self.grades = {}

    def add_submission(self, student_id, content):
        if student_id in self.submissions:
            return False
        self.submissions[student_id] = {
            'content': content,
            'submitted_at': datetime.now().isoformat()
        }
        return True

class Grade:
    def __init__(self, grade_id, student_id, subject, value, date, teacher_id):
        if not 1 <= value <= 5:
            raise ValueError("Grade must be between 1 and 5")
        self.grade_id = grade_id
        self.student_id = student_id
        self.subject = subject
        self.value = value
        self.date = date
        self.teacher_id = teacher_id

    def update_grade(self, new_value):
        if not 1 <= new_value <= 5:
            raise ValueError("Grade must be between 1 and 5")
        self.value = new_value
        self.date = datetime.now().isoformat()


  


class DataExporter:

    @staticmethod
    def export_to_xlsx(users:list, filename='eduplatform_data.xlsx'):
        try:
            wb=Workbook()
            wb_users=wb.active
            wb_users.title='Users'
            wb_users.append(['ID', 'Full-Name', 'Email', 'Role', 'Created at'])
            for user in users:
                wb_users.append([
                    user._id,
                    user._full_name,
                    user._email,
                    user.role.value,
                    user._created_at
                ])

            #Student sheet
            wb_students=wb.create_sheet('Students')
            wb_students.append(["ID", 'Grade', 'Average Grade'])
            for user in users:
                if isinstance(user, Student):
                    avg=user.calculate_average_grade()
                    wb_students.append([user._id, user.grade, avg])
            # Teacher sheet
            wb_teachers=wb.create_sheet('Teachers') 
            wb_teachers.append(['ID', 'Subjects'])
            for user in users:
                if isinstance(user, Teacher):
                    wb_teachers.append([user._id,', '.join(user.subjects)])     

            # Parent sheet
            wb_parents=wb.create_sheet('Parents')
            wb_parents.append(['ID', 'Child ID'])
            for user in users:
                if isinstance(user, Parent):
                    wb_parents.append([user._id, user.child_id])

            #save file
            wb.save(filename)
            print(f'Date exported to {filename}') 
            return True
        except Exception as e:
            print(f'Error exporting file to  XLSX: {e}')    
            return False
        
    @staticmethod
    def export_to_csv(users: list, filename='EduPlatform'):    
        try:
            with open(f'{filename}_users.csv', 'w', newline= '') as f:
                writer=csv.writer(f)
                writer.writerow(['ID', "FUll-NAME", 'Email', "Role", 'Created ad'])
                for user in users:
                    writer.writerow([
                        user._id,
                        user._full_name,
                        user._email,
                        user.role.value,
                        user._created_at
                    ])
            # Students csv
            with open(f'{filename}_students.csv', 'w', newline= '') as f:
                writer=csv.writer(f)
                writer.writerow(['ID', 'Grade', 'Average grade'])
                for user in users:
                    if isinstance(user, Student):
                        avg=user.calculate_average_grade()
                        writer.writerow([user._id, avg])


            # teacher csv
            with open(f'{filename}_teacher.csv', 'w', newline= '') as f:
                writer=csv.writer(f)
                writer.writerow(['ID', 'Subject'])
                for user in users:
                    if isinstance(user, Teacher):
                        writer.writerow([user._id, ','.join(user.subjects)])


            # parent csv
            with open(f'{filename}_parent.csv', 'w', newline= '') as f:
                writer=csv.writer(f)
                writer.writerow(['ID', 'Child ID '])
                for user in users:
                    if isinstance(user, Parent):
                        writer.writerow([user._id, user.child_id])  

            print(f'Data exported to CSV with prefix {filename}') 
            return True
        except Exception     as e:
            print(f'Error exporting file to  XLSX: {e}')    
            return False

    @staticmethod 
    def export_to_sql(users: list, filename:str ='Eduplatform_data.sql'):
        try:
            with open(filename, 'w') as f:
                f.write(""" 
--User table
Create table USERS(
    ID int primary key,
    Full_Name varchar(20) NOT NULL,
    Email varchar(20) Unique NOT NULL,
    created_at DATETIME not null,
    password_hash VARCHAR(50) NOt null   
);
  

-- Create Students table
CREATE TABLE Students (
    user_id INT PRIMARY KEY,
    grade VARCHAR(20),
    FOREIGN KEY (user_id) REFERENCES Users(id)
);

-- Create Teachers table
CREATE TABLE Teachers (
    user_id INT PRIMARY KEY,
    subjects TEXT,
    FOREIGN KEY (user_id) REFERENCES Users(id)
);

-- Create Parents table
CREATE TABLE Parents (
    user_id INT PRIMARY KEY,
    children TEXT,
    FOREIGN KEY (user_id) REFERENCES Users(id)
);

-- Create Notifications table
CREATE TABLE Notifications (
    id INT PRIMARY KEY,
    recipient_id INT NOT NULL,
    sender_id INT NOT NULL,
    message TEXT NOT NULL,
    created_at DATETIME NOT NULL,
    is_read BIT DEFAULT 0,
    FOREIGN KEY (recipient_id) REFERENCES Users(id),
    FOREIGN KEY (sender_id) REFERENCES Users(id)
);

-- Create Assignments table
CREATE TABLE Assignments (
    id INT PRIMARY KEY,
    title VARCHAR(100) NOT NULL,
    description TEXT,
    deadline DATETIME,
    subject VARCHAR(50),
    teacher_id INT NOT NULL,
    class_id VARCHAR(20),
    FOREIGN KEY (teacher_id) REFERENCES Teachers(user_id)
);

-- Create Grades table
CREATE TABLE Grades (
    id INT PRIMARY KEY,
    student_id INT NOT NULL,
    assignment_id INT NOT NULL,
    grade INT CHECK (grade BETWEEN 1 AND 5),
    comment TEXT,
    FOREIGN KEY (student_id) REFERENCES Students(user_id),
    FOREIGN KEY (assignment_id) REFERENCES Assignments(id)
);                                                                                                                                      )                                     
""")
            
            # inset data
            f.write('\nInsert into\n')
            for user in users:
                # In export_to_sql():
                f.write(f"""
INSERT INTO USERS(id, Full_Name, Email, created_at, password_hash)
VALUES ({user._id}, '{user._full_name}', '{user._email}', '{user._created_at}', '{user._password_hash}');
""")   
                f.write("\n-- Insert Students\n")
                for user in users:
                    if isinstance(user, Student):
                        f.write(f"""
INSERT INTO Students (user_id, grade)
VALUES ({user._id}, '{user.grade}');
""")

                f.write("\n-- Insert Teachers\n")
                for user in users:
                    if isinstance(user, Teacher):
                        f.write(f"""
INSERT INTO Teachers (user_id, subjects)
VALUES ({user._id}, '{", ".join(user.subjects)}');
""")
                        
                f.write("\n-- Insert Parents\n")
                for user in users:
                    if isinstance(user, Parent):
                        children = ",".join(str(cid) for cid in user.children.keys())
                        f.write(f"""
INSERT INTO Parents (user_id, children)
VALUES ({user._id}, '{children}');
""")       
            print(f'DATA exported to sql {filename}')
            return True
        except Exception as e:
            print(f"Error generating SQL: {e}")
            return False    
        
    
    @staticmethod
    def export_all(users: list):
        """Export data to all formats with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = f"eduplatform_export_{timestamp}"
        
        # Log the export
        with open("export_log.txt", "a") as log:
            log.write(f"{datetime.now()}: Export started\n")
            
            xlsx_success = DataExporter.export_to_xlsx(users, f"{prefix}.xlsx")
            log.write(f"{datetime.now()}: XLSX export {'succeeded' if xlsx_success else 'failed'}\n")
            
            csv_success = DataExporter.export_to_csv(users, prefix)
            log.write(f"{datetime.now()}: CSV export {'succeeded' if csv_success else 'failed'}\n")
            
            sql_success = DataExporter.export_to_sql(users, f"{prefix}.sql")
            log.write(f"{datetime.now()}: SQL export {'succeeded' if sql_success else 'failed'}\n")
            
            log.write(f"{datetime.now()}: Export completed\n\n")  


        

                        




              

                    





# ================= CLI SECTION =================
def main():
    print("Welcome to EduPlatform CLI")
    users = []
    admin = Admin("Samadjon Sayfullayev", "samadjonsayfullayev3106@admin.com", "123")
    admin_users = [admin]
    users.append(admin)
    parent_users = []
    try:
        while True:
            print("\nChoose your role: [1] Admin, [2] Teacher, [3] Student, [4] Parent, [5] Exit")
            print("Registered Users:")
            for u in users:
                print(f"{u._full_name} - ID: {u._id} - Role: {u.role.value}")

            role = input("Enter role number: ")

            if role == '1':
                try:
                    uid = int(input("ID: "))
                except ValueError:
                    print("Invalid ID format.")
                    continue
                pwd = input("Password: ")
                admin = next((u for u in admin_users if u._id == uid), None)
                if not admin or admin._password_hash != hashlib.sha256(pwd.encode()).hexdigest():
                    print("Invalid credentials or not registered.")
                    continue
                    
                print("Admin options: [1] Add User, [2] Remove User, [3] View Profile, [4] Update Profile [5] Link Parent-Child, [6] Manage Schedules")
                action = input("Choose action: ")
                if action == '1':
                    name = input("Full Name: ")
                    email = input("Email: ")
                    pwd = input("Password: ")
                    print("User type: [1] Student, [2] Teacher, [3] Parent")
                    user_type = input("Choose: ")
                    if user_type == '1':
                        grade = input("Grade: ")
                        new_user = Student(name, email, pwd, grade)
                    elif user_type == '2':
                        subjects = input("Subjects: ").split(',')
                        new_user = Teacher(name, email, pwd, subjects)
                    elif user_type == '3':
                        new_user = Parent(name, email, pwd)
                        parent_users.append(new_user)
                    else:
                        print("Invalid user type")
                        continue
                    if not admin_users:
                        admin = Admin("Samadjon Sayfullayev", "samadjonsayfullayev3106@admin.com", "admin")
                        admin_users.append(admin)
                    admin_users[0].add_user(new_user, users)
                    users.append(new_user)
                    print("User added by Admin.")
                    print(f"User ID: {new_user._id}")
                    admin.add_user(new_user, users)  # Add to admin's user dictionary
                    
                    
                elif action == '2':
                    uid = int(input("Enter user id to remove: "))
                    if admin_users and admin_users[0].remove_user(uid):
                        print("User removed.")
                    else:
                        print("User not found.")
                elif action == '3':
                    print("\nProfile Information:")
                    for key, value in admin.get_profile().items():
                        print(f"{key.replace('_', ' ').title()}: {value}")
                elif action == '4':
                    print("Update options: [1] Full Name, [2] Email, [3] Password")
                    update_choice = input("What to update: ")
                    if update_choice == '1':
                        new_name = input("New full name: ")
                        admin.update_profile(full_name=new_name)
                    elif update_choice == '2':
                        new_email = input("New email: ")
                        admin.update_profile(email=new_email)
                    elif update_choice == '3':
                        new_pwd = input("New password: ")
                        admin.update_profile(password=new_pwd)
                    else:
                        print("Invalid option")
                    print("Profile updated successfully")
                elif action=='5':
                    parent_id=int(input("Enter parent ID: "))
                    student_id=int(input('Enter student ID: '))  
                    admin.link_parent_child(parent_id,student_id, users) 
                elif action=='6':
                    print('[1] Create schedule [2] add lesson [3] remove Lesson [4] view schedule') 
                    schedule_action=input('Enter your action: ') 
                    if schedule_action=='1':
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        admin.create_schedule(class_id, day)
                    elif schedule_action=='2':
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        schedule= admin.get_schedule(class_id, day)
                        if not schedule:
                            print('Schedule not found')
                            continue
                        time=input('Enter time( e.g 9.00-10.00): ')
                        subject=input('Enter name of subject: ')
                        teacher_id=int(input('Enter teacher ID: '))
                        schedule.add_lesson(time, subject, teacher_id, users)
                    elif schedule_action=='3':
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        schedule= admin.get_schedule(class_id, day, users)
                        if not schedule:
                            print('Schedule not found')
                            continue
                        time=input('ENter time to remove: ')
                        schedule.remove_lesson(time)
                    elif schedule_action=='4':
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        schedule= admin.get_schedule(class_id, day)
                        if schedule:
                            schedule.view_schedule(users)
                        else:
                            print("Schedule not found")
                    else: 
                        print('Invalid action') 
                        continue       
                        

            elif role == '2':
                try:
                    uid = int(input("ID: "))
                except ValueError:
                    print("Invalid ID format.")
                    continue
                pwd = input("Password: ")
                teacher = next((u for u in users if isinstance(u, Teacher) and u._id == uid), None)
                if not teacher or teacher._password_hash != hashlib.sha256(pwd.encode()).hexdigest():
                    print("Invalid credentials or not registered.")
                    continue
                
                print("[1] Create Assignment, [2] Grade Assignment, [3] Send notification to Parent, [4] View Profile, [5] Update Profile [6] View schedule")
                choice = input("Choose: ")
                
                if choice == '1':
                    title = input("Title: ")
                    description = input("Description: ")
                    deadline = input("Deadline: ")
                    subject = input("Subject: ")
                    class_id = input("Class ID: ")
                    aid = teacher.create_assignment(title, description, deadline, subject, class_id)
                    print(f"Assignment created with ID: {aid}")
                
                elif choice == '2':
                    aid = int(input("Assignment ID: "))
                    sid = int(input("Student ID: "))
                    grade = int(input("Grade (1-5): "))
                    comment = input("Comment: ")
                    success = teacher.grade_assignment(aid, sid, grade, comment, users)
                    
                    if success:
                        print("Graded successfully.")
                        print("Do you want to send notification to student's parents? [1] Yes [2] No")
                        try:
                            yes_no = int(input("Choice: "))
                            if yes_no == 1:
                                # Find student to get their name for the message
                                student = next((s for s in users if isinstance(s, Student) and s._id == sid), None)
                                if student:
                                    message = (f"Your child {student._full_name} received grade {grade} "
                                            f"for assignment {aid} with comment: {comment}")
                                    # Send to all parents of this student
                                    parents = [p for p in users if isinstance(p, Parent) and sid in p.children]
                                    if parents:
                                        for parent in parents:
                                            notif_id = len(parent._notifications) + 1
                                            notification = Notification(notif_id, message, teacher._id, parent._id)
                                            if notification.send(users):
                                                print(f"Notification sent to parent {parent._full_name}")
                                            else:
                                                print(f"Failed to send to parent {parent._full_name}")
                                    else:
                                        print("No parents found for this student.")
                                else:
                                    print("Student not found.")
                            elif yes_no == 2:
                                print("Notification not sent.")
                            else:
                                print("Invalid choice. Notification not sent.")
                        except ValueError:
                            print("Invalid input. Notification not sent.")
                    else:
                        print("Failed to grade.")
                
                elif choice == '3':  # Direct message to parent option
                    sid = int(input("Student ID: "))
                    message = input("Message to parent: ")
                    teacher.send_to_parent(sid, message, users)
                
                elif choice == '4':
                    print("\nProfile Information:")
                    for key, value in teacher.get_profile().items():
                        print(f"{key.replace('_', ' ').title()}: {value}")
                    print(f"Subjects: {', '.join(teacher.subjects)}")
                
                elif choice == '5':
                    print("Update options: [1] Full Name, [2] Email, [3] Password, [4] Subjects")
                    update_choice = input("What to update: ")
                    if update_choice == '1':
                        new_name = input("New full name: ")
                        teacher.update_profile(full_name=new_name)
                    elif update_choice == '2':
                        new_email = input("New email: ")
                        teacher.update_profile(email=new_email)
                    elif update_choice == '3':
                        new_pwd = input("New password: ")
                        teacher.update_profile(password=new_pwd)
                    elif update_choice == '4':
                        new_subjects = input("New subjects (comma separated): ").split(',')
                        teacher.subjects = [s.strip() for s in new_subjects]
                    else:
                        print("Invalid option")
                    print("Profile updated successfully")
                elif choice=='6':
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        schedule= admin.get_schedule(class_id, day)
                        if schedule:
                            schedule.view_schedule(users)
                        else:
                            print("Schedule not found")
                            
                        
            elif role == '3':
                try:
                    uid = int(input("ID: "))
                except ValueError:
                    print("Invalid ID format.")
                    continue
                pwd = input("Password: ")
                student = next((u for u in users if isinstance(u, Student) and u._id == uid), None)
                if not student or student._password_hash != hashlib.sha256(pwd.encode()).hexdigest():
                    print("Invalid credentials or not registered.")
                    continue
                print("[1] Submit Assignment, [2] View Grades, [3] Calculate Average, [4] View Profile, [5] Update Profile [6] View Schedule")
                choice = input("Choose: ")
                if choice == '1':
                    aid = int(input("Assignment ID: "))
                    content = input("Content: ")
                    success = student.submit_assignment(aid, content)
                    print("Submitted." if success else "Already submitted.")
                elif choice == '2':
                    student.view_grades()
                elif choice == '3':
                    print("Average:", student.calculate_average_grade())
                elif choice == '4':
                    print("\nProfile Information:")
                    for key, value in student.get_profile().items():
                        print(f"{key.replace('_', ' ').title()}: {value}")
                    print(f"Grade: {student.grade}")
                elif choice == '5':
                    print("Update options: [1] Full Name, [2] Email, [3] Password, [4] Grade")
                    update_choice = input("What to update: ")
                    if update_choice == '1':
                        new_name = input("New full name: ")
                        student.update_profile(full_name=new_name)
                    elif update_choice == '2':
                        new_email = input("New email: ")
                        student.update_profile(email=new_email)
                    elif update_choice == '3':
                        new_pwd = input("New password: ")
                        student.update_profile(password=new_pwd)
                    elif update_choice == '4':
                        new_grade = input("New grade: ")
                        student.grade = new_grade
                    else:
                        print("Invalid option")
                    print("Profile updated successfully")
                elif choice=='6':    
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        schedule= admin.get_schedule(class_id, day)
                        if schedule:
                            schedule.view_schedule(users)
                        else:
                            print("Schedule not found")
            elif role == '4':
                try:
                    uid = int(input("ID: "))
                except ValueError:
                    print("Invalid ID format.")
                    continue
                pwd = input("Password: ")
                parent = next((p for p in parent_users if p._id == uid), None)
                if not parent or parent._password_hash != hashlib.sha256(pwd.encode()).hexdigest():
                    print("Invalid credentials.")
                    continue
                print("[1] View Child Grades, [2] View Child Assignments, [3] View Profile, [4] Update Profile , [5] Notification [6]View Schedule ")
                action = input("Choose: ")
                if action in ['1', '2']:
                    cid = int(input("Enter Child ID: "))
                    if action == '1':
                        parent.view_child_grades(cid, users)
                    elif action == '2':
                        parent.view_child_assignments(cid, users)
                elif action == '3':
                    print("\nProfile Information:")
                    for key, value in parent.get_profile().items():
                        print(f"{key.replace('_', ' ').title()}: {value}")
                    print(f"Children IDs: {', '.join(str(cid) for cid in parent.children.keys())}")
                elif action == '4':
                    print("Update options: [1] Full Name, [2] Email, [3] Password")
                    update_choice = input("What to update: ")
                    if update_choice == '1':
                        new_name = input("New full name: ")
                        parent.update_profile(full_name=new_name)
                    elif update_choice == '2':
                        new_email = input("New email: ")
                        parent.update_profile(email=new_email)
                    elif update_choice == '3':
                        new_pwd = input("New password: ")
                        parent.update_profile(password=new_pwd)
                    else:
                        print("Invalid option")
                    print("Profile updated successfully")
                
    
                elif action == '5':  # View notifications
                    print("\n=== Notifications ===")
                    # In CLI parent section (action == '5'):
                    for notification in parent._notifications:
                        print(notification)  # Remove users parameter
                        print("-" * 30)
                elif action=='6':
                        class_id=int(input('Enter class ID: '))
                        day=input('Day of week: ')
                        schedule= admin.get_schedule(class_id, day)
                        if schedule:
                            schedule.view_schedule(users)
                        else:
                            print("Schedule not found")     

            elif role == '5':
                    print("Exporting data before exit...")
                    DataExporter.export_all(users)
                    print("Goodbye!")
                    break 

            else:
                print("Invalid role")
    except Exception as e:
        print(f"An error occurred: {e}")
        print("Attempting to export data before exiting...")
        DataExporter.export_all(users)
        print("Emergency export completed. Goodbye!")
if __name__ == "__main__":
    main()